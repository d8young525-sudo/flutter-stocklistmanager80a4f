#!/usr/bin/env python3
"""
Firebase 가격표 업로드 스크립트
Excel PRICECHART 시트에서 MODEL, MY, PRICE를 읽어서 Firestore에 업로드
"""

import sys
import openpyxl
from pathlib import Path

# Firebase Admin SDK 임포트
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
    print("✅ firebase-admin 패키지 로드 성공")
except ImportError as e:
    print(f"❌ firebase-admin 패키지를 찾을 수 없습니다: {e}")
    print("📦 설치 명령어: pip install firebase-admin==7.1.0")
    sys.exit(1)

# Firebase Admin SDK 초기화
def initialize_firebase(sdk_path):
    """Firebase Admin SDK 초기화"""
    try:
        if not Path(sdk_path).exists():
            print(f"❌ Firebase Admin SDK 파일을 찾을 수 없습니다: {sdk_path}")
            sys.exit(1)
        
        cred = credentials.Certificate(sdk_path)
        firebase_admin.initialize_app(cred)
        print("✅ Firebase Admin SDK 초기화 완료")
        return firestore.client()
    except Exception as e:
        print(f"❌ Firebase 초기화 오류: {e}")
        sys.exit(1)

def parse_price_excel(excel_path):
    """
    Excel PRICECHART 시트에서 가격 데이터 파싱
    
    Returns:
        list: [{'model': str, 'my': str, 'price': int}, ...]
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # PRICECHART 시트 찾기
        if 'PRICECHART' not in wb.sheetnames:
            print(f"❌ 'PRICECHART' 시트를 찾을 수 없습니다.")
            print(f"   사용 가능한 시트: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        
        sheet = wb['PRICECHART']
        print(f"✅ 'PRICECHART' 시트 로드 완료")
        
        # 헤더 행 찾기 (MODEL, MY, PRICE 컬럼)
        header_row = None
        model_col = None
        my_col = None
        price_col = None
        
        for row_idx in range(1, min(21, sheet.max_row + 1)):  # 처음 20행 검색
            for cell in sheet[row_idx]:
                if cell.value:
                    cell_value = str(cell.value).strip().upper()
                    if 'MODEL' in cell_value:
                        model_col = cell.column
                    elif cell_value in ['MY', 'MODEL YEAR']:
                        my_col = cell.column
                    elif 'PRICE' in cell_value or '가격' in cell_value:
                        price_col = cell.column
            
            if model_col and my_col and price_col:
                header_row = row_idx
                break
        
        if not (model_col and my_col and price_col):
            print(f"❌ 필수 컬럼을 찾을 수 없습니다 (MODEL, MY, PRICE)")
            sys.exit(1)
        
        print(f"📋 헤더 행: {header_row}, MODEL: {model_col}, MY: {my_col}, PRICE: {price_col}")
        
        # 데이터 파싱
        prices = []
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            model_cell = sheet.cell(row=row_idx, column=model_col)
            my_cell = sheet.cell(row=row_idx, column=my_col)
            price_cell = sheet.cell(row=row_idx, column=price_col)
            
            model = str(model_cell.value).strip() if model_cell.value else ""
            my = str(my_cell.value).strip() if my_cell.value else ""
            
            # 가격 파싱
            price = 0
            if price_cell.value:
                try:
                    if isinstance(price_cell.value, (int, float)):
                        price = int(price_cell.value)
                    else:
                        # 문자열에서 숫자만 추출
                        price_str = str(price_cell.value).replace(',', '').replace(' ', '')
                        price = int(''.join(filter(str.isdigit, price_str)))
                except:
                    pass
            
            # 유효한 데이터만 추가
            if model and my and price > 0:
                prices.append({
                    'model': model,
                    'my': my,
                    'price': price
                })
        
        print(f"✅ {len(prices)}개의 가격 데이터 파싱 완료")
        return prices
        
    except Exception as e:
        print(f"❌ Excel 파싱 오류: {e}")
        sys.exit(1)

def upload_prices_to_firestore(db, prices):
    """
    Firestore에 가격 데이터 업로드
    
    Collection: prices
    Document ID: {model}_{my}
    Fields: model, my, price
    """
    try:
        batch = db.batch()
        prices_ref = db.collection('prices')
        
        print(f"\n🔄 Firestore 업로드 시작...")
        
        for idx, price_data in enumerate(prices, 1):
            doc_id = f"{price_data['model']}_{price_data['my']}"
            doc_ref = prices_ref.document(doc_id)
            
            batch.set(doc_ref, price_data, merge=True)
            
            # 진행 상황 출력
            if idx % 10 == 0 or idx == len(prices):
                print(f"   {idx}/{len(prices)} 업로드 중...")
        
        # Batch commit
        batch.commit()
        print(f"✅ {len(prices)}개의 가격 데이터 Firestore 업로드 완료!")
        
        # 업로드된 데이터 샘플 출력
        print(f"\n📊 업로드된 데이터 샘플:")
        for i, price_data in enumerate(prices[:5], 1):
            formatted_price = f"{price_data['price']:,}원"
            print(f"   {i}. {price_data['model']} ({price_data['my']}) - {formatted_price}")
        
        if len(prices) > 5:
            print(f"   ... 외 {len(prices) - 5}개")
        
    except Exception as e:
        print(f"❌ Firestore 업로드 오류: {e}")
        sys.exit(1)

def main():
    """메인 실행 함수"""
    print("="*60)
    print("🔥 Firebase 가격표 업로드 스크립트")
    print("="*60)
    
    # Excel 파일 경로 확인
    if len(sys.argv) < 2:
        print("\n사용법:")
        print("  python upload_prices_to_firebase.py <가격표_엑셀_파일.xlsx>")
        print("\n예시:")
        print("  python upload_prices_to_firebase.py 가격표.xlsx")
        print("\n⚠️ Firebase Admin SDK 파일이 같은 폴더에 있어야 합니다!")
        print("   파일명: firebase-admin-sdk.json")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not Path(excel_path).exists():
        print(f"❌ 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    
    print(f"📁 Excel 파일: {excel_path}\n")
    
    # Firebase Admin SDK 파일 찾기 (같은 폴더)
    sdk_path = Path(__file__).parent / "firebase-admin-sdk.json"
    if not sdk_path.exists():
        print(f"❌ Firebase Admin SDK 파일을 찾을 수 없습니다: {sdk_path}")
        print("💡 firebase-admin-sdk.json 파일을 스크립트와 같은 폴더에 넣어주세요!")
        sys.exit(1)
    
    # Firebase 초기화
    db = initialize_firebase(str(sdk_path))
    
    # Excel 파싱
    prices = parse_price_excel(excel_path)
    
    # Firestore 업로드
    upload_prices_to_firestore(db, prices)
    
    print("\n" + "="*60)
    print("✅ 모든 작업 완료!")
    print("="*60)
    print("\n💡 Flutter 앱에서 자동으로 가격표를 다운로드합니다.")

if __name__ == "__main__":
    main()
