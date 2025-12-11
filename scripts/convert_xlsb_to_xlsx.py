#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
.xlsb 파일을 .xlsx 파일로 변환하는 스크립트
Flask API 서버에서 사용 가능
"""

import sys
import io
from pyxlsb import open_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def convert_xlsb_to_xlsx_bytes(xlsb_bytes):
    """
    .xlsb 바이너리 데이터를 .xlsx 바이너리 데이터로 변환
    
    Args:
        xlsb_bytes: .xlsb 파일의 바이너리 데이터 (bytes)
        
    Returns:
        .xlsx 파일의 바이너리 데이터 (bytes)
    """
    try:
        # 바이너리 데이터를 BytesIO로 변환하여 pyxlsb로 읽기
        xlsb_stream = io.BytesIO(xlsb_bytes)
        
        # openpyxl Workbook 생성
        wb_out = Workbook()
        wb_out.remove(wb_out.active)  # 기본 시트 제거
        
        # .xlsb 파일 읽기
        with open_workbook(xlsb_stream) as wb_in:
            for sheet_name in wb_in.sheets:
                # 새 시트 생성
                ws_out = wb_out.create_sheet(title=sheet_name)
                
                # 시트 데이터 읽기
                with wb_in.get_sheet(sheet_name) as sheet:
                    row_idx = 1
                    for row in sheet.rows():
                        col_idx = 1
                        for cell in row:
                            # 셀 값이 None이 아닌 경우만 복사
                            if cell.v is not None:
                                ws_out.cell(row=row_idx, column=col_idx, value=cell.v)
                            col_idx += 1
                        row_idx += 1
        
        # .xlsx 파일을 BytesIO로 저장
        output_stream = io.BytesIO()
        wb_out.save(output_stream)
        output_stream.seek(0)
        
        return output_stream.read()
        
    except Exception as e:
        raise Exception(f"변환 중 오류 발생: {str(e)}")


def convert_xlsb_file_to_xlsx_file(xlsb_path, xlsx_path):
    """
    .xlsb 파일을 .xlsx 파일로 변환 (파일 경로 사용)
    
    Args:
        xlsb_path: 입력 .xlsb 파일 경로
        xlsx_path: 출력 .xlsx 파일 경로
    """
    try:
        with open(xlsb_path, 'rb') as f:
            xlsb_bytes = f.read()
        
        xlsx_bytes = convert_xlsb_to_xlsx_bytes(xlsb_bytes)
        
        with open(xlsx_path, 'wb') as f:
            f.write(xlsx_bytes)
            
        print(f"✅ 변환 완료: {xlsb_path} → {xlsx_path}")
        
    except Exception as e:
        print(f"❌ 변환 실패: {str(e)}")
        sys.exit(1)


# CLI 사용 예시
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("사용법: python3 convert_xlsb_to_xlsx.py <입력.xlsb> <출력.xlsx>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    convert_xlsb_file_to_xlsx_file(input_file, output_file)
