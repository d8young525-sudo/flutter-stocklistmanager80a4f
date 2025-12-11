#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Firebase Cloud Functions용 .xlsb → .xlsx 변환 스크립트
stdin에서 .xlsb 데이터를 읽고, stdout으로 .xlsx 데이터를 출력
"""

import sys
import io
from pyxlsb import open_workbook
from openpyxl import Workbook

def convert_xlsb_to_xlsx(xlsb_bytes):
    """
    .xlsb 바이너리 데이터를 .xlsx 바이너리 데이터로 변환
    """
    try:
        # 바이너리 데이터를 BytesIO로 변환
        xlsb_stream = io.BytesIO(xlsb_bytes)
        
        # openpyxl Workbook 생성
        wb_out = Workbook()
        wb_out.remove(wb_out.active)  # 기본 시트 제거
        
        # .xlsb 파일 읽기
        with open_workbook(xlsb_stream) as wb_in:
            for sheet_name in wb_in.sheets:
                # 새 시트 생성
                ws_out = wb_out.create_sheet(title=sheet_name)
                
                # 시트 데이터 읽기
                with wb_in.get_sheet(sheet_name) as sheet:
                    row_idx = 1
                    for row in sheet.rows():
                        col_idx = 1
                        for cell in row:
                            # 셀 값이 None이 아닌 경우만 복사
                            if cell.v is not None:
                                ws_out.cell(row=row_idx, column=col_idx, value=cell.v)
                            col_idx += 1
                        row_idx += 1
        
        # .xlsx 파일을 BytesIO로 저장
        output_stream = io.BytesIO()
        wb_out.save(output_stream)
        output_stream.seek(0)
        
        return output_stream.read()
        
    except Exception as e:
        sys.stderr.write(f"Conversion error: {str(e)}\n")
        sys.exit(1)


if __name__ == "__main__":
    # stdin에서 .xlsb 바이너리 데이터 읽기
    xlsb_data = sys.stdin.buffer.read()
    
    if not xlsb_data:
        sys.stderr.write("No input data received\n")
        sys.exit(1)
    
    # 변환 실행
    xlsx_data = convert_xlsb_to_xlsx(xlsb_data)
    
    # stdout으로 .xlsx 바이너리 데이터 출력
    sys.stdout.buffer.write(xlsx_data)
    sys.stdout.buffer.flush()
